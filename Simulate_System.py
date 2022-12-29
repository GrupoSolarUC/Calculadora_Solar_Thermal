# -*- coding: utf-8 -*-
"""
Created on Wed Dec  7 15:30:01 2022

@author: adria
"""

import os
import tkinter
from timeit import default_timer
import openpyxl
import csv
import sys
from scipy.integrate import quad
from scipy.interpolate import interp1d
from math import pi, sin, cos
import numpy as np
import pvlib as pv
from pvlib.solarposition import get_solarposition
import pandas as pd
import matplotlib
from matplotlib import pyplot as plt
from mpl_toolkits.axes_grid1 import make_axes_locatable

from Files.Solar_Position import azimuth_function, zenith_function
from Files.Class_Solar_Field import Solar_Field
from Files.Class_Storage import Storage_Tank
from Files.Class_Heat_Exchanger import Heat_Exchanger
from Files.Class_Properties import Properties
 
plt.ioff()
plt.rc('xtick', labelsize=14)
plt.rc('ytick', labelsize=14)

def raise_error(error_message):
    window=tkinter.Tk()
    window.attributes('-topmost',1)
    window.withdraw()
    tkinter.messagebox.showerror('Error',error_message)
    window.deiconify()
    window.attributes("-topmost", True)
    window.destroy()
    window.quit()


    
def Heat_Map(Curve, title, filename):
    if len(Curve)%365 != 0:
        print('Número de elementos en el vector de resultados no es válido')
        return
    time_steps_per_day = int( len( Curve )/365 )
    yticks = [ 0, int(time_steps_per_day/6), int(time_steps_per_day/3),
               int(time_steps_per_day/2), int(time_steps_per_day*2/3),
               int(time_steps_per_day*5/6), time_steps_per_day]
    ylabels = [0,4,8,12,16,20,24]
    matrix = np.reshape(Curve, ( 365, time_steps_per_day ) )
    matrix = np.transpose(matrix)
    fig = plt.figure(figsize = (8,12), dpi = 200)
    ax = plt.gca()
    im = ax.matshow(matrix, cmap=plt.cm.jet, origin= 'lower')
    divider = make_axes_locatable(ax)
    cax = divider.append_axes("right", size="5%", pad=0.2)
    plt.colorbar(im, cax=cax)
    ax.set_yticks(yticks, ylabels)
    ax.xaxis.set_ticks_position('bottom')
    ax.xaxis.set_label_position('bottom')
    ax.set_ylabel('Hora del día', fontsize = 14)
    ax.set_xlabel('Día del año', fontsize = 14)
    ax.set_title(title, fontsize = 16, loc = 'left')
    plt.savefig(filename+'.jpg', bbox_inches='tight')

start = default_timer()

## Define parameters file name
params_file_name = 'Introducir_Parametros.xlsx'

## Load parameters file
try:
    params_file = openpyxl.load_workbook(params_file_name)
## If the parameters file cannot be found, a notification is yielded
except FileNotFoundError:
    raise_error("No se encuentra el archivo 'Introducir_Parametros.xlsx'. Este archivo debe estar en la misma carpeta donde se encuentra el Script 'Simulate_System.py'")
    sys.exit()
## If the parameters file is open, it cannot be imported and a notification is yielded
except PermissionError:
    raise_error("Cierre el archivo 'Introducir_Parametros.xlsx'")
    sys.exit()
params = params_file.active


## IMPORT PARAMETERS FROM PARAMETERS FILE
## THE PARAMETERS ARE ANALYZED TO CHECK THEIR VALIDITY
## ALSO, SOME UNITS ARE CHANGED IN ORDER TO USE HOURS AS TIME UNIT

## Import simulation number
simulation_number = params.cell(row = 1, column = 2).value
if not type(simulation_number) == int:
    raise_error('Número de la simulación (fila 1) debe ser un número entero')
    params_file.close()
    sys.exit()
    
if os.path.exists('Resultados_Simulacion_'+str(simulation_number)):
    raise_error("Ya existe una carpeta con el nombre 'Resultados_Simulacion_"+str(simulation_number)+"'. Por favor eliminarla, moverla, cambiarle el nombre o escoger otro número para la simulación actual (fila 1 en el archivo de parámetros)")
    params_file.close()
    sys.exit()

## Import name of the climate data file
climate_data_file_name = str(params.cell(row = 2, column = 2).value)
if not climate_data_file_name[len(climate_data_file_name)-4:] == '.csv':
    climate_data_file_name = climate_data_file_name + '.csv'
    
## Code of the collector
collector_name = str(params.cell(row = 3, column = 2).value)

## Area of each collector
coll_A = params.cell(row = 4, column = 2).value
if not ((type(coll_A) == int or type(coll_A) == float) and coll_A > 0):
    raise_error('El área de los colectores (fila 4) debe ser un número mayor a cero (expresando el área de un colector en m^2)')
    params_file.close()
    sys.exit()

## Optical efficiency of the collectors
coll_n0 = params.cell(row = 5, column = 2).value
if not ((type(coll_n0) == int or type(coll_n0) == float) and coll_n0 > 0):
    raise_error('La eficiencia óptica del colector (fila 5) debe ser un número mayor a cero')
    params_file.close()
    sys.exit()
if coll_n0 > 1:
    if coll_n0 > 100:
        raise_error('La eficiencia óptica del colector (fila 5) no puede ser mayor a 100')
        params_file.close()
        sys.exit()
    coll_n0 = coll_n0/100

## Linear and quadratic heat loss coefficient of the collectors
coll_a1_watt = params.cell(row = 6, column = 2).value
coll_a2_watt = params.cell(row = 7, column = 2).value
if not ((type(coll_a1_watt) == float or type(coll_a1_watt) == int) and (type(coll_a2_watt) == float or type(coll_a2_watt) == int) and coll_a1_watt >= 0 and coll_a2_watt >= 0):
    raise_error('Coeficientes de pérdidas lineales y cuadráticas de los colectores (filas 6 y 7) deben ser valores mayores o iguales a cero (expresando los valores en W/(m^2.K) y W/(m^2.K^2) respectivamente')
    params_file.close()
    sys.exit()
coll_a1 = coll_a1_watt*3.6
coll_a2 = coll_a2_watt*3.6

## Import type of IAM data (monoaxial, biaxial)
IAM_type = params.cell(row = 8, column = 2).value

## Import IAM values
if IAM_type in ['Monoaxial', 'monoaxial', 'MONOAXIAL']:
    IAM_dict = {}
    col = 3
    while True:
        angle = params.cell(row = 9, column = col).value
        if angle == None:
            break
        IAM_value = params.cell(row = 10, column = col).value
        no_value = params.cell(row = 11, column = col).value
        if not ((type(angle) == int or type(angle) == float) and angle >= 0 and angle <= 90):
            raise_error('Ángulos especificados en la fila 9 deben estar en el rango [0, 90] (se permiten ambos límites del intervalo)')
            params_file.close()
            sys.exit()
        if no_value != None:
            raise_error("Eliminar valores de la fila 11, o cambiar parámetro 'Tipo de datos de IAM' (fila 6) a 'Biaxial'")
            params_file.close()
            sys.exit()
        if not (type(IAM_value) == float or type(IAM_value) == int):
            raise_error('Para cada ángulo especificado debe haber un valor de IAM en la fila 10')
            params_file.close()
            sys.exit()
        if not IAM_value >= 0:
            raise_error('Valores de IAM (fila 10) deben ser iguales o mayores a cero')
            params_file.close()
            sys.exit()
        IAM_dict[angle] = IAM_value
        col = col + 1
elif IAM_type in ['Biaxial', 'biaxial', 'BIAXIAL']:
    IAM_dict = {'K_l': {}, 'K_t': {}}
    col = 3
    while True:
        angle = params.cell(row = 9, column = col).value
        if angle == None:
            break
        Kl_value = params.cell(row = 10, column = col).value
        Kt_value = params.cell(row = 11, column = col).value
        if not ((type(angle) == int or type(angle) == float) and angle >= 0 and angle <= 90):
            raise_error('Ángulos especificados en la fila 9 deben estar en el rango [0, 90] (se permiten ambos límites del intervalo)')
            params_file.close()
            sys.exit()
        if not ((type(Kl_value) == float or type(Kl_value) == int) and (type(Kt_value) == float or type(Kt_value) == int)):
            raise_error("Para cada ángulo especificado debe haber un valor de IAM transversal y longitudinal (filas 10 y 11). Si los datos de IAM del colector sólo tienen una fila de datos, cambiar parámetro 'Tipo de datos de IAM' (fila 6) a 'Monoaxial'")
            params_file.close()
            sys.exit()
        if not (Kl_value >= 0 and Kt_value >= 0):
            raise_error('Valores de IAM (filas 10 y 11) deben ser iguales o mayores a cero')
            params_file.close()
            sys.exit()
        IAM_dict['K_l'][angle] = Kl_value
        IAM_dict['K_t'][angle] = Kt_value
        col = col + 1
## If no IAM data was imported because the 'Type of IAM data' parameter has an invalid name, a notification is yielded
else:
    raise_error("Valor de 'Tipo de datos de IAM' (fila 8) no válido")
    params_file.close()
    sys.exit()

## Collector test fluid
coll_test_fluid = params.cell(row = 12, column = 2).value
if type(coll_test_fluid) == int or type(coll_test_fluid) == float:
    coll_test_glycol_percentage = coll_test_fluid
    if coll_test_glycol_percentage < 0:
        raise_error("Porcentaje de glicol en el fluido de testeo (fila 12) debe ser un valor mayor a cero (o en su defecto una de las palabras: 'Agua' o 'Glicol')")
        params_file.close()
        sys.exit()
    if coll_test_glycol_percentage > 100:
        raise_error("Porcentaje de glicol en el fluido de testeo (fila 12) debe ser menor a 100 (o en su defecto una de las palabras: 'Agua' o 'Glicol')")
        params_file.close()
        sys.exit()
    if coll_test_glycol_percentage > 0 and coll_test_glycol_percentage < 1:
        coll_test_glycol_percentage = int(np.round(coll_test_glycol_percentage*100, 0))
    if not coll_test_glycol_percentage in [0,10,15,20,25,30,40,50,60,100]:
        raise_error('Porcentaje de glicol en el fluido de testeo (fila 12) no válido. Los valores permitidos son: 0, 10, 15, 20, 25, 30, 40, 50, 60, 100')
        params_file.close()
        sys.exit()
else:
    if not coll_test_fluid in ['agua', 'Agua', 'AGUA', 'water', 'Water', 'WATER', 'glicol', 'Glicol', 'GLICOL', 'glycol', 'Glycol', 'GLYCOL']:
        raise_error('Fluido de prueba del colector (fila 12) no válido (ver nota)')
        params_file.close()
        sys.exit()
    if coll_test_fluid in ['agua', 'Agua', 'AGUA', 'water', 'Water', 'WATER']:
        coll_test_glycol_percentage = 0
    if coll_test_fluid in ['glicol', 'Glicol', 'GLICOL', 'glycol', 'Glycol', 'GLYCOL']:
        coll_test_glycol_percentage = 40

## Collector test flow (per area unit)
coll_test_flow_per_m2_seg = params.cell(row = 13, column = 2).value
if not ((type(coll_test_flow_per_m2_seg) or type(coll_test_flow_per_m2_seg) == int) and coll_test_flow_per_m2_seg > 0):
    raise_error('Flujo de prueba (por unidad de área) del colector (fila 13) debe ser un número mayor a cero (expresando el flujo de prueba en kg/(m^2*s)')
    params_file.close()
    sys.exit()
coll_test_flow_per_m2 = coll_test_flow_per_m2_seg*3600

## Number of collector rows and number of collectors per row
coll_rows = params.cell(row = 14, column = 2).value
colls_per_row = params.cell(row = 15, column = 2).value
if not (type(coll_rows) == int and type(colls_per_row) == int and coll_rows > 0 and colls_per_row > 0):
    raise_error('Número de filas de colectores y número de colectores por fila (filas 14 y 15) deben ser valores enteros mayores a cero')
    params_file.close()
    sys.exit()
    
## Azimuth and tilt of the collectors
coll_tilt = params.cell(row = 16, column = 2).value
coll_azimuth = params.cell(row = 17, column = 2).value
if not ((type(coll_tilt) == int or type(coll_tilt) == float) and coll_tilt >= 0 and coll_tilt <= 90):
    raise_error('Inclinación de los colectores (fila 16) debe ser un valor un valor numérico entre 0 y 90 (expresando el ángulo de inclinación en grados; los límites del rango están permitidos)')
    params_file.close()
    sys.exit()
if not ((type(coll_azimuth) == int or type(coll_azimuth) == float) and coll_azimuth >= 0 and coll_azimuth < 360):
    raise_error('Azimut de los colectores (fila 17) debe ser un valor un valor numérico entre 0 y 360 (expresando el ángulo azimutal en grados). El valor puede ser cero pero no 360. El ángulo se mide desde la dirección norte hacia el este.')
    params_file.close()
    sys.exit()

## Tank volume
tank_volume_L = params.cell(row = 18, column = 2).value
if not ((type(tank_volume_L) == float or type(tank_volume_L) == int) and tank_volume_L > 0):
    raise_error('Volumen del tanque (fila 18) debe ser un valor mayor a cero (expresando el volumen del tanque en litros)')
    params_file.close()
    sys.exit()
tank_volume = tank_volume_L/1000

## Tank's aspect ratio
tank_AR = params.cell(row = 19, column = 2).value
if not ((type(tank_AR) == float or type(tank_AR) == int) and tank_AR >= 0.5 and tank_AR <= 4 ):
    raise_error('Relación de aspecto del tanque (fila 19) debe ser un valor mayor o igual a 0.5 y menor o igual a 4')
    params_file.close()
    sys.exit()

## Glycol percentage in the solar field
glycol_percentage = params.cell(row = 20, column = 2).value
if not ((type(glycol_percentage) == int or type(glycol_percentage) == float) and glycol_percentage >= 0):
    raise_error('Porcentaje de glicol (fila 20) debe ser un valor numérico mayor a cero')
    params_file.close()
    sys.exit()
if glycol_percentage > 100:
    raise_error('Porcentaje de glicol (fila 20) no puede ser mayor a 100')
    params_file.close()
    sys.exit()
if glycol_percentage > 0 and glycol_percentage < 1:
    glycol_percentage = int(np.round(glycol_percentage*100,0))
if not glycol_percentage in [0,10,15,20,25,30,40,50,60,100]:
    raise_error('Valor de porcentaje de glicol (fila 20) no permitido. Los valores posibles son: 0, 10, 15, 20, 25, 30, 40, 50, 60, 100')
    params_file.close()
    sys.exit()

## Flow in the solar field
solar_field_flow_min = params.cell(row = 21, column = 2).value
if not ((type(solar_field_flow_min) == int or type(solar_field_flow_min) == float) and solar_field_flow_min > 0):
    raise_error('Flujo en el campo solar (fila 21) debe ser un valor mayor a cero (expresando el flujo en L/min)')
    params_file.close()
    sys.exit()
solar_field_flow = solar_field_flow_min*60

## Type of demand profile (possible values = 1, 2 or 3)
demand_type = params.cell(row = 22, column = 2).value
if not demand_type in [1,2,3]:
    raise_error("Valor de 'Tipo de perfil de demanda' (fila 22) inválido. Este parámetro debe ser uno de los siguientes valores: 1, 2, 3. Ver nota en la celda correspondiente para más detalles sobre cada opción")
    params_file.close()
    sys.exit()

## Demand values (one per hour of one day)
demand_points_min = [ params.cell(row = 24, column = i).value for i in range(3, 27) ]
for demand in demand_points_min:
    if not ((type(demand) == float or type(demand) == int) and demand >= 0):
        raise_error('Datos de flujo demandado (fila 24) deben ser números mayores o iguales a cero (expresando el fujo en kg/s)')
        params_file.close()
        sys.exit()
demand_points = [ 60*demand for demand in demand_points_min ]

## Setpoint temperature
setpoint = params.cell(row = 25, column = 2).value
if not ((type(setpoint) == float or type(setpoint) == int) and setpoint >= 30):
    raise_error('Setpoint del calentador auxiliar (fila 25) debe ser un valor numérico igual o mayor a 30 (expresando la temperatura en °C)')
    params_file.close()
    sys.exit()

## Yearly cost of heating water with conventional energy sources
yearly_cost = params.cell(row = 26, column = 2).value
if yearly_cost != None and not ((type(yearly_cost) == int or type(yearly_cost) == float) and yearly_cost>= 0):
    raise_error("Costo anual de calentar agua por medios convencionales (fila 26) debe ser un valor mayor a cero (también puede dejarse la celda vacía)")
    params_file.close()
    sys.exit()
        
params_file.close()

## Parameters whose values are not imposed by the user and are thus assumed.

tank_top_loss_coeff = 5
tank_edge_loss_coeff = 5
tank_bottom_loss_coeff = 5
tank_flow = solar_field_flow
HX_eff = 0.8
tank_nodes = 10
albedo = 0.3
solar_field_pressure = 5
tank_pressure = 2
time_step = 0.1

## The minimum irradiation level necessary to operate the solar field is computed as the minimum irradiation that compensates the thermal losses, assuming that there is a difference of 15K between the ambient temperature and the temperature of the flow that enters the solar field, plus a 10% margin.

rad_min = (coll_a1*15 + coll_a2*15**2)/coll_n0
rad_min = rad_min*1.1

## Define demand_func(t), a function that receives the hour of the year (from 0 to 8760) and returns the demanded flow at that moment. The definition of the function depends on the parameter "Tipo de perfil de demanda" specified by the user on the Excel file

if demand_type == 1:
    demand_points = demand_points*365
    demand_points.append(demand_points[0])
    demand_time = range(8761)
    interp_demand = interp1d(demand_time, demand_points)
    def demand_func(t):
        return float(interp_demand(t))
if demand_type == 2:
    def demand_func(t):
        daytime_int = int(t%24)
        return demand_points[daytime_int]
if demand_type == 3:
    demand_points.append(demand_points[0])
    def demand_func(t):
        daytime = t%24
        daytime_int = int(t%24)
        if daytime - daytime_int < 0.5:
            return demand_points[daytime_int]
        else:
            return demand_points[daytime_int + 1]

## Import climate data from the file specified by the user

try:
    climate_data_file = open(climate_data_file_name)
    climate_data = csv.reader(climate_data_file)
except FileNotFoundError:
    raise_error("No se encuentra el archivo de datos climáticos. Este archivo debe estar en la misma carpeta donde se encuentra el Script 'Simulate_System.py'")
    params_file.close()
    sys.exit()
climate_data = [ row for row in climate_data ]
climate_data_file.close()

## Extract relevant data from climate data

latitude = float(climate_data[11][1])
longitude = float(climate_data[12][1])
DNI = [ 3.6*float(climate_data[i][8]) for i in [8801] + list(range(42, 8801)) ]
GHI = np.array([ 3.6*float(climate_data[i][5]) for i in [8801] + list(range(42, 8801)) ])
DHI = np.array([ 3.6*float(climate_data[i][3]) for i in [8801] + list(range(42, 8801)) ])
temp = [ float(climate_data[i][9]) for i in [8801] + list(range(42, 8801)) ]

## Define a function to approximate the mains water temperature. The method is taken from "TOWARDS DEVELOPMENT OF AN ALGORITHM FOR MAINS WATER TEMPERATURE" from Burch and Christensen

temp_january = np.mean(temp[:744])
temp_february = np.mean(temp[744:1416])
temp_march = np.mean(temp[1416:2160])
temp_april = np.mean(temp[2160:2880])
temp_may = np.mean(temp[2880:3624])
temp_june = np.mean(temp[3624:4344])
temp_july = np.mean(temp[4344:5088])
temp_august = np.mean(temp[5088:5832])
temp_september = np.mean(temp[5832:6552])
temp_october = np.mean(temp[6552:7296])
temp_november = np.mean(temp[7296:8016])
temp_december = np.mean(temp[8016:8760])
T_month_list = [temp_january,
                temp_february,
                temp_march,
                temp_april,
                temp_may,
                temp_june,
                temp_july,
                temp_august,
                temp_september,
                temp_october,
                temp_november,
                temp_december ]
T_amb_ann = np.mean(temp)
delta_T_amb = ( max(T_month_list) - min(T_month_list) )/2
delta_T_offset = 3.3333333
T_ref = 6.6666667
K1 = 0.4
K2 = 0.018
K3 = 35*pi/180
K4 = -3.1416e-4
delta_T_mains = (K1 + K2*(T_amb_ann - T_ref))*delta_T_amb
phi_lag = K3 + K4*(T_amb_ann - T_ref)
phi_amb = (104.8 + 180)*pi/180
T_mains_avg = T_amb_ann + delta_T_offset
def T_mains_func(t):
    return T_mains_avg + delta_T_mains*sin(2*pi*t/8760 - phi_lag - phi_amb)

## Define a function that yields the time of day according to the time zone
## This function supposes that the year begins at monday

if latitude < 49:
    def corrected_time(t):
        if t >= 24*97 and t < 24*251:
            return t - 1
        else:
            return t
else:
    def corrected_time(t):
            return t

## Define functions to yield radiation data as a function of time

diffuse = pv.irradiance.isotropic(coll_tilt, DHI)
diff_ground = pv.irradiance.get_ground_diffuse(coll_tilt, GHI, albedo)

diffuse = diffuse.tolist()
diff_ground = diff_ground.tolist()

DNI.append(DNI[0])
diffuse.append(diffuse[0])
diff_ground.append(diff_ground[0])
temp.append(temp[0])

time = range(8761)

diffuse_func = interp1d(time, diffuse)
diff_ground_func = interp1d(time, diff_ground)
DNI_func = interp1d(time, DNI)
T_amb_func = interp1d(time, temp)

## Define functions to yield solar position as a function of time

january_year = climate_data[42][0][:4]
february_year = climate_data[786][0][:4]
march_year = climate_data[1458][0][:4]
april_year = climate_data[2202][0][:4]
may_year = climate_data[2922][0][:4]
june_year = climate_data[3666][0][:4]
july_year = climate_data[4386][0][:4]
august_year = climate_data[5130][0][:4]
september_year = climate_data[5874][0][:4]
october_year = climate_data[6594][0][:4]
november_year = climate_data[7338][0][:4]
december_year = climate_data[8058][0][:4]


time_january = pd.date_range(january_year + '-01-01', january_year + '-02-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_february = pd.date_range(february_year + '-02-01', february_year + '-03-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_march = pd.date_range(march_year + '-03-01', march_year + '-04-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_april = pd.date_range(april_year + '-04-01', april_year + '-05-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_may = pd.date_range(may_year + '-05-01', may_year + '-06-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_june = pd.date_range(june_year + '-06-01', june_year + '-07-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_july = pd.date_range(july_year + '-07-01', july_year + '-08-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_august = pd.date_range(august_year + '-08-01', august_year + '-09-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_september = pd.date_range(september_year + '-09-01', september_year + '-10-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_october = pd.date_range(october_year + '-10-01', october_year + '-11-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_november = pd.date_range(november_year + '-11-01', november_year + '-12-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')
time_december = pd.date_range(december_year + '-12-01', str(int(december_year) + 1) + '-01-01', inclusive='left', freq='min', tz = 'Etc/GMT+3')

if len(time_february) == 41760:
    time_february = pd.date_range(february_year + '-02-01', february_year + '-02-29', inclusive='left', freq='min', tz = 'Etc/GMT+3')

solpos_january = get_solarposition(time_january, latitude, longitude)
azimuth_january = solpos_january['azimuth']
zenith_january = solpos_january['zenith']

solpos_february = get_solarposition(time_february, latitude, longitude)
azimuth_february = solpos_february['azimuth']
zenith_february = solpos_february['zenith']

solpos_march = get_solarposition(time_march, latitude, longitude)
azimuth_march = solpos_march['azimuth']
zenith_march = solpos_march['zenith']

solpos_april = get_solarposition(time_april, latitude, longitude)
azimuth_april = solpos_april['azimuth']
zenith_april = solpos_april['zenith']

solpos_may = get_solarposition(time_may, latitude, longitude)
azimuth_may = solpos_may['azimuth']
zenith_may = solpos_may['zenith']

solpos_june = get_solarposition(time_june, latitude, longitude)
azimuth_june = solpos_june['azimuth']
zenith_june = solpos_june['zenith']

solpos_july = get_solarposition(time_july, latitude, longitude)
azimuth_july = solpos_july['azimuth']
zenith_july = solpos_july['zenith']

solpos_august = get_solarposition(time_august, latitude, longitude)
azimuth_august = solpos_august['azimuth']
zenith_august = solpos_august['zenith']

solpos_september = get_solarposition(time_september, latitude, longitude)
azimuth_september = solpos_september['azimuth']
zenith_september = solpos_september['zenith']

solpos_october = get_solarposition(time_october, latitude, longitude)
azimuth_october = solpos_october['azimuth']
zenith_october = solpos_october['zenith']

solpos_november = get_solarposition(time_november, latitude, longitude)
azimuth_november = solpos_november['azimuth']
zenith_november = solpos_november['zenith']

solpos_december = get_solarposition(time_december, latitude, longitude)
azimuth_december = solpos_december['azimuth']
zenith_december = solpos_december['zenith']

solar_azimuth_list = (azimuth_january.tolist() +
                      azimuth_february.tolist() +
                      azimuth_march.tolist() +
                      azimuth_april.tolist() +
                      azimuth_may.tolist() +
                      azimuth_june.tolist() +
                      azimuth_july.tolist() +
                      azimuth_august.tolist() +
                      azimuth_september.tolist() +
                      azimuth_october.tolist() +
                      azimuth_november.tolist() +
                      azimuth_december.tolist() )

solar_zenith_list = (zenith_january.tolist() +
                     zenith_february.tolist() +
                     zenith_march.tolist() +
                     zenith_april.tolist() +
                     zenith_may.tolist() +
                     zenith_june.tolist() +
                     zenith_july.tolist() +
                     zenith_august.tolist() +
                     zenith_september.tolist() +
                     zenith_october.tolist() +
                     zenith_november.tolist() +
                     zenith_december.tolist() )

compute_azimuth = azimuth_function(solar_azimuth_list)
compute_zenith = zenith_function(solar_zenith_list)

## Define type of solar collector (monoaxial IAM or biaxial IAM)

if IAM_type in ['Monoaxial', 'monoaxial', 'MONOAXIAL']:
    coll_type = 'Monoaxial'
if IAM_type in ['Biaxial', 'biaxial', 'BIAXIAL']:
    coll_type = 'Biaxial'
    
## Define fluid properties

motherFolder = os.getcwd()
os.chdir('Files')

PropsCollTest = Properties(coll_test_glycol_percentage, solar_field_pressure)
PropsField = Properties(glycol_percentage, solar_field_pressure)
PropsTank = Properties(0, tank_pressure)

os.chdir(motherFolder)

## Define devices of the system

tank_initial_temperatures = {i: 20 for i in range(1, tank_nodes + 1)}
Field = Solar_Field(coll_type, coll_A, coll_n0, coll_a1, coll_a2,
                    IAM_dict, coll_test_flow_per_m2, PropsCollTest,
                    coll_tilt, coll_azimuth, coll_rows, colls_per_row, PropsField)
Tank = Storage_Tank(tank_volume, tank_AR, tank_top_loss_coeff,
                    tank_edge_loss_coeff, tank_bottom_loss_coeff,
                    tank_nodes, 1, 10, 10, 1, PropsTank,
                    tank_initial_temperatures)
HX = Heat_Exchanger(HX_eff, PropsField, PropsTank)

## Begin simulation 10 days before January 1, in order to avoid the effect of the initial conditions inside the storage tank. The simulation results of these first 10 days are not recorded.

t = 8760 - 24*10
previous_enthalpies = {'h_in_solar_field': PropsField.T_to_h(40), 'h_in_HX_cold': PropsTank.T_to_h(40)}
while t < 8760:
    diff = float(diffuse_func(t))
    diff_ground = float(diff_ground_func(t))
    DNIrr = float(DNI_func(t))
    T_amb = float(T_amb_func(t))
    T_mains = T_mains_func(t)
    h_mains = PropsTank.T_to_h(T_mains)
    demanded_flow = demand_func(corrected_time(t))
    aoi, longi, trans = Field.incidence_angles(float(compute_zenith(t)), float(compute_azimuth(t)))
    if aoi == None or longi == None or trans == None:
        operation = False
    else:
        beam = DNIrr*cos(aoi)
        if diff + diff_ground + beam > rad_min:
            operation = True
        else:
            operation = False
    if operation:
        it = 0
        while True:
            Field_outputs = Field.compute_outputs(solar_field_flow, previous_enthalpies['h_in_solar_field'],
                                                  h_mains, beam, diff, diff_ground,
                                                  aoi, trans, longi, T_amb)
            HX_outputs = HX.compute_outputs(solar_field_flow, Field_outputs['h_out'],
                                            tank_flow, previous_enthalpies['h_in_HX_cold'], h_mains)
            Tank_outputs = Tank.compute_outputs(tank_flow, HX_outputs['h_out_load'],
                                                demanded_flow, h_mains,
                                                T_amb, time_step)
            it = it + 1
            if ( abs(HX_outputs['h_out_source'] - previous_enthalpies['h_in_solar_field'])/previous_enthalpies['h_in_solar_field'] < 1e-6 and
                  abs(Tank_outputs['outlet_1_h'] - previous_enthalpies['h_in_HX_cold'])/previous_enthalpies['h_in_HX_cold'] < 1e-6):
                maxIt = False
                break
            if it == 100:
                maxIt = True
                break
            previous_enthalpies = {'h_in_solar_field': HX_outputs['h_out_source'], 'h_in_HX_cold': Tank_outputs['outlet_1_h']}
        Tank.update_temperature()
    else:
        Tank_outputs = Tank.compute_outputs(0, h_mains,
                                            demanded_flow, h_mains,
                                            T_amb, time_step)
        Tank.update_temperature()
    t = t + time_step
    t = np.round(t, 1)
    
## Begin simulation from January 1 to December 31.

Result = []
t = 0
while t < 8760:
    diff = float(diffuse_func(t))
    diff_ground = float(diff_ground_func(t))
    DNIrr = float(DNI_func(t))
    T_amb = float(T_amb_func(t))
    T_mains = T_mains_func(t)
    h_mains = PropsTank.T_to_h(T_mains)
    demanded_flow = demand_func(corrected_time(t))
    solar_zenith = float(compute_zenith(t))
    solar_azimuth = float(compute_azimuth(t))
    aoi, longi, trans = Field.incidence_angles(solar_zenith, solar_azimuth)
    if aoi == None or longi == None or trans == None:
        operation = False
    else:
        beam = DNIrr*cos(aoi)
        if diff + diff_ground + beam > rad_min:
            operation = True
        else:
            operation = False
    if operation:
        it = 0
        while True:
            Field_outputs = Field.compute_outputs(solar_field_flow, previous_enthalpies['h_in_solar_field'],
                                                  h_mains, beam, diff, diff_ground,
                                                  aoi, trans, longi, T_amb)
            HX_outputs = HX.compute_outputs(solar_field_flow, Field_outputs['h_out'],
                                            tank_flow, previous_enthalpies['h_in_HX_cold'], h_mains)
            Tank_outputs = Tank.compute_outputs(tank_flow, HX_outputs['h_out_load'],
                                                demanded_flow, h_mains,
                                                T_amb, time_step)
            it = it + 1
            if ( abs(HX_outputs['h_out_source'] - previous_enthalpies['h_in_solar_field'])/previous_enthalpies['h_in_solar_field'] < 1e-6 and
                  abs(Tank_outputs['outlet_1_h'] - previous_enthalpies['h_in_HX_cold'])/previous_enthalpies['h_in_HX_cold'] < 1e-6):
                maxIt = False
                break
            if it == 100:
                maxIt = True
                break
            previous_enthalpies = {'h_in_solar_field': HX_outputs['h_out_source'], 'h_in_HX_cold': Tank_outputs['outlet_1_h']}
        Tank.update_temperature()
        Result.append({'t': t,
                       'operation': operation,
                       'T_mains': T_mains,
                       'demanded_flow': demanded_flow,
                       'demanded_flow_temp': PropsTank.h_to_T(Tank_outputs['outlet_2_h']),
                       'T_in_solar_field': PropsField.h_to_T(HX_outputs['h_out_source']),
                       'T_out_solar_field': PropsField.h_to_T(Field_outputs['h_out']),
                       'solar_field_Q_useful': Field_outputs['Q_useful'],
                       'solar_field_Q_waste': Field_outputs['Q_waste'],
                       'HX_Q_useful': HX_outputs['Q_useful'],
                       'HX_Q_waste': HX_outputs['Q_waste'],
                       'T_in_HX_cold': PropsTank.h_to_T(Tank_outputs['outlet_1_h']),
                       'T_out_HX_cold': PropsTank.h_to_T(HX_outputs['h_out_load']),
                       'Heat_load': Tank_outputs['Q_demand'],
                       'Heat_loss': Tank_outputs['Q_loss'],
                       'beam_rad': beam,
                       'diff_rad': diff,
                       'grnd_rad': diff_ground,
                       'aoi': aoi,
                       'longitudinal_angle': longi,
                       'transverse_angle': trans,
                       'solar_zenith': solar_zenith,
                       'solar_azimuth': solar_azimuth,
                       'maxIt_system': maxIt,
                       'maxIt_tank': Tank_outputs['maxIt'],
                       'maxIt_solar': Field_outputs['maxIt'] } )
    else:
        Tank_outputs = Tank.compute_outputs(0, h_mains,
                                            demanded_flow, h_mains,
                                            T_amb, time_step)
        Tank.update_temperature()
        Result.append({'t': t,
                       'operation': operation,
                       'T_mains': T_mains,
                       'demanded_flow': demanded_flow,
                       'demanded_flow_temp': PropsTank.h_to_T(Tank_outputs['outlet_2_h']),
                       'T_in_solar_field': None,
                       'T_out_solar_field': None,
                       'solar_field_Q_useful': 0,
                       'solar_field_Q_waste': 0,
                       'HX_Q_useful': 0,
                       'HX_Q_waste': 0,
                       'T_in_HX_cold': None,
                       'T_out_HX_cold': None,
                       'Heat_load': Tank_outputs['Q_demand'],
                       'Heat_loss': Tank_outputs['Q_loss'],
                       'beam_rad': beam,
                       'diff_rad': diff,
                       'grnd_rad': diff_ground,
                       'aoi': aoi,
                       'longitudinal_angle': longi,
                       'transverse_angle': trans,
                       'solar_zenith': solar_zenith,
                       'solar_azimuth': solar_azimuth,
                       'maxIt_system': False,
                       'maxIt_tank': Tank_outputs['maxIt'],
                       'maxIt_solar': False } )
    t = t + time_step
    t = np.round(t, 1)
Result = {key: [ Result[i][key] for i in range(len(Result)) ] for key in Result[0] }

os.makedirs('Resultados_Simulacion_'+str(simulation_number))
os.chdir('Resultados_Simulacion_'+str(simulation_number))

daily_consumption = quad(demand_func, 0, 24, limit = 10000)[0]
demand_curve = [ demand_func(t)/60 for t in np.linspace(0,24,10000) ]
plt.figure(figsize = (15,8), dpi = 200)
plt.plot(np.linspace(0,24,10000), demand_curve)
plt.xlabel('Hora del día', fontsize = 14)
plt.ylabel('Consumo [L/min]', fontsize = 14)
plt.title('Consumo diario total: '+str(np.round(daily_consumption,1))+'L', fontsize = 16, loc = 'left')
plt.savefig('Consumo_Diario.jpg', bbox_inches='tight')

Power_needed = [ 4.184*Result['demanded_flow'][i]*(setpoint - Result['T_mains'][i])/3.6 for i in range(len(Result['demanded_flow'])) ]
Preheat_Power = [ 4.184*Result['demanded_flow'][i]*(min([ setpoint, Result['demanded_flow_temp'][i] ]) - Result['T_mains'][i])/3.6 for i in range(len(Result['demanded_flow'])) ]
Total_Solar_Fraction = sum(Preheat_Power)/sum(Power_needed)
instantaneous_solar_frac = []
for i in range(len(Power_needed)):
    if Power_needed[i] > 0:
        instantaneous_solar_frac.append(Preheat_Power[i]/Power_needed[i])
    else:
        instantaneous_solar_frac.append(np.nan)
instantaneous_solar_frac_masked = np.ma.array(instantaneous_solar_frac, mask = np.isnan(instantaneous_solar_frac))
cmap = matplotlib.cm.jet
cmap.set_bad('white',1.)
Q_useful = [ Result['HX_Q_useful'][i]/3.6 for i in range(len(Result['solar_field_Q_useful'])) ]
Q_waste = [ (Result['HX_Q_waste'][i] + Result['solar_field_Q_waste'][i])/3.6 for i in range(len(Result['solar_field_Q_useful'])) ]
maximum_solar_frac = np.nanmax(instantaneous_solar_frac)
minimum_solar_frac = np.nanmin(instantaneous_solar_frac)

excess_temp_counter = 0
demand_time_counter = 0
for i in range(len(Result['demanded_flow_temp'])):
    if Result['demanded_flow'][i] > 0 and Result['demanded_flow_temp'][i] > setpoint:
        excess_temp_counter = excess_temp_counter + 1
    if Result['demanded_flow'][i] > 0:
        demand_time_counter = demand_time_counter + 1
        
total_demand_time = np.round(demand_time_counter*time_step,1)

if excess_temp_counter > 0:
    excess_temp_time = np.round(excess_temp_counter*time_step,1)
    excess_temp_time_percentage = np.round(100*excess_temp_time/total_demand_time, 2)


Heat_Map(Result['demanded_flow_temp'], 'Temperatura del agua saliendo del tanque [°C]', 'T_out_Tank')
Heat_Map(instantaneous_solar_frac, 'Fracción solar', 'Solar_Fraction')
Heat_Map(Q_useful, 'Calor útil aportado por campo solar [W]', 'Q_useful')


Q_waste_counter = 0
for i in range(len(Q_waste)):
    if Q_waste[i] > 0:
        Q_waste_counter = Q_waste_counter + 1
        
if Q_waste_counter > 0:
    Heat_Map(Q_useful, 'Calor purgado en forma de vapor [W]', 'Q_waste')
Q_waste_time = np.round(time_step*Q_waste_counter,1)

operation_counter = 0
for i in range(len(Result['operation'])):
    if Result['operation'][i]:
        operation_counter = operation_counter + 1
operation_time = np.round(time_step*operation_counter,1)
Q_waste_time_percentage = np.round(100*Q_waste_time/operation_time,2)
Q_waste_percentage = np.round(100*sum(Q_waste)/sum(Q_useful),2)


Parameters_file = open('Parametros_Simulacion.txt','w')
Parameters_file.write('Número de la simulación: '+str(simulation_number)+'\n')
Parameters_file.write('Nombre de archivo con datos climáticos: '+climate_data_file_name+'\n')
if collector_name != None:
    Parameters_file.write('Nombre, modelo o código del colector: '+collector_name+'\n')
Parameters_file.write('Área de cada colector: '+str(coll_A)+'m^2\n')
Parameters_file.write('Intersección de la curva de eficiencia de los colectores: '+str(coll_n0)+'\n')
Parameters_file.write('Coeficiente de pérdidas térmicas lineales de los colectores: '+str(coll_a1_watt)+'W/(m^2 K)\n')
Parameters_file.write('Coeficiente de pérdidas térmicas cuadráticas de los colectores: '+str(coll_a2_watt)+'W/(m^2 K^2)\n')
Parameters_file.write('Tipo de datos de IAM: '+IAM_type+'\n')
Parameters_file.write('\n')
Parameters_file.write('Valores de IAM:\n')
if IAM_type in ['Monoaxial', 'monoaxial', 'MONOAXIAL']:
    for angle in IAM_dict:
        Parameters_file.write(' - '+str(angle)+'°: '+str(IAM_dict[angle])+'\n')
if IAM_type in ['Biaxial', 'biaxial', 'BIAXIAL']:
    for angle in IAM_dict['K_l']:
        Parameters_file.write(' - '+str(angle)+'°: IAM longitudinal = '+str(IAM_dict['K_l'][angle])+' ; IAM transversal = '+str(IAM_dict['K_t'][angle])+'\n')
Parameters_file.write('\n')
Parameters_file.write('Porcentaje de glicol en el fluido de prueba del colector: '+str(coll_test_glycol_percentage)+'%\n')
Parameters_file.write('Flujo de prueba, por unidad de área, del colector: '+str(coll_test_flow_per_m2_seg)+'kg/(m^2 s)\n')
Parameters_file.write('Número de filas de colectores: '+str(coll_rows)+'\n')
Parameters_file.write('Número de colectores por fila: '+str(colls_per_row)+'\n')
Parameters_file.write('Inclinación de los colectores respecto al plano horizontal: '+str(coll_tilt)+'°\n')
Parameters_file.write('Azimut de los colectores: '+str(coll_azimuth)+'°\n')
Parameters_file.write('Volumen del tanque: '+str(tank_volume_L)+'L\n')
Parameters_file.write('Relación de aspecto del tanque: '+str(tank_AR)+'\n')
Parameters_file.write('Porcentaje de glicol en el campo solar: '+str(glycol_percentage)+'%\n')
Parameters_file.write('Flujo en el campo solar: '+str(solar_field_flow_min)+'kg/min\n')
Parameters_file.write('Tipo de perfil de demanda: '+str(demand_type)+'\n')
Parameters_file.write('\n')
time_list = ['00:00', '01:00', '02:00', '03:00', '04:00', '05:00', '06:00', '07:00', '08:00', '09:00', '10:00', '11:00', 
             '12:00', '13:00', '14:00', '15:00', '16:00', '17:00', '18:00', '19:00', '20:00', '21:00', '22:00', '23:00', ]
Parameters_file.write('Perfil de consumo: \n')
for i in range(len(time_list)):
    Parameters_file.write(' - '+time_list[i]+': '+str(demand_points_min[i])+' L/min\n')
Parameters_file.write('\n')
Parameters_file.write('Setpoint del calentador auxiliar: '+str(setpoint)+'°C\n')
if yearly_cost != None:
    Parameters_file.write('Costo anual de calentar agua por medios convencionales: $'+str(yearly_cost)+'\n')
Parameters_file.close()

Results_file = open('Resultados + ReadMe.txt','w')
Results_file.write('Curva de demanda de agua especificada produce una demanda total diaria de '+str(np.round(daily_consumption,1))+'L.\n')
Results_file.write('El calor total (útil; no purgado en forma de vapor) aportado por el campo solar durante el año fue de '+str(np.round(sum(Q_useful)*time_step/1000,2))+' kWh.\n')
Results_file.write('Horario en los mapas ajustado al horario UTC-3 (horario de verano en Chile). Es normal que se observe un desplazamiento en la demanda en invierno debido a los cambios de hora. Se asume que éstos se realizan los días 7 de abril y 8 de septiembre a las 24:00 horas.\n')
Results_file.write('En el mapa de fracción solar, las zonas blancas corresponden a horarios en los que no existe flujo de demandado; esto hace que se indefina la fracción solar.\n')
if Q_waste_counter == 0:
    Results_file.write('No se ha generado un mapa de calor purgado debido a que no se ha generado vapor en el campo solar ni en el intercambiador de calor durante todo el año simulado.\n')
Results_file.write('Fracción solar anual: '+str(Total_Solar_Fraction)+'\n')
if yearly_cost != None:
    Results_file.write('Ahorro anual estimado en fuentes de energía convencionales: '+str(np.round(Total_Solar_Fraction*yearly_cost,2))+'\n')
Results_file.write('Fracción solar mínima: '+str(np.round(minimum_solar_frac,2))+'\n')
Results_file.write('Fracción solar máxima: '+str(np.round(maximum_solar_frac,2))+'\n')
if excess_temp_counter > 0:
    Results_file.write('La temperatura del agua demandada saliendo del tanque excedió la temperatura deseada (setpoint del calentador auxiliar) durante '+str(excess_temp_time)+' horas; esto representa un '+str(excess_temp_time_percentage)+'% del tiempo total durante el cual se demandó agua caliente.\n')
    Results_file.write('El agua saliendo del tanque alcanzó una temperatura máxima de '+str(np.round(max(Result['demanded_flow_temp']),1))+'°C\n')
    Results_file.write('Considere reducir el tamaño del campo solar si la temperatura alcanzada fuera excesiva, o si el porcentaje del tiempo durante el cual se excedió la temperatura deseada fuera considerable.\n')
else:
    Results_file.write('La temperatura del agua demandada saliendo del tanque no llegó a exceder la temperatura deseada (setpoint del calentador auxiliar).\n')
    Results_file.write('El agua saliendo del tanque alcanzó una temperatura máxima de '+str(np.round(max(Result['demanded_flow_temp']),1))+'°C.\n')
if Q_waste_counter > 0:
    Results_file.write('Se ha porducido y expulsado vapor durante un tiempo total de '+str(Q_waste_time)+'; esto representa un '+str(Q_waste_time_percentage)+'% del tiempo total de operación del campo solar.\n')
    Results_file.write("El calor usado para producir vapor (y por lo tanto 'perdido') representa un "+str(Q_waste_percentage)+'% del calor útil aportado por el sistema solar térmico durante el año simulado. Si existe un margen para subir la temperatura en el tanque sin pasar el setpoint del calentador, considere aumentar el flujo en el campo solar. Si no existe dicho margen, considere disminuir el área del campo solar.\n')
Results_file.close()

Result2 = {}

Result2['t'] = Result['t']
Result2['T_mains'] = Result['T_mains']
Result2['operation'] = Result['operation']
Result2['demanded_flow'] = [ Result['demanded_flow'][i]/60 for i in range(len(Result['demanded_flow'])) ]
Result2['demanded_flow_temp'] = Result['demanded_flow_temp']
Result2['T_in_solar_field'] = Result['T_in_solar_field']
Result2['T_out_solar_field'] = Result['T_out_solar_field']
Result2['solar_field_Q_useful'] = [ Result['solar_field_Q_useful'][i]/3.6 for i in range(len(Result['solar_field_Q_useful'])) ]
Result2['solar_field_Q_waste'] = [ Result['solar_field_Q_waste'][i]/3.6 for i in range(len(Result['solar_field_Q_waste'])) ]
Result2['HX_Q_useful'] = [ Result['HX_Q_useful'][i]/3.6 for i in range(len(Result['HX_Q_useful'])) ]
Result2['HX_Q_waste'] = [ Result['HX_Q_waste'][i]/3.6 for i in range(len(Result['HX_Q_waste'])) ]
Result2['T_in_HX_load_side'] = Result['T_in_HX_cold']
Result2['T_out_HX_load_side'] = Result['T_out_HX_cold']
Result2['Heat_delivered_to_demanded_flow'] = [ Result['Heat_load'][i]/3.6 for i in range(len(Result['Heat_load'])) ]
Result2['Heat_losses_through_tank_walls'] = [ Result['Heat_loss'][i]/3.6 for i in range(len(Result['Heat_loss'])) ]
Result2['beam_rad'] = [ Result['beam_rad'][i]/3.6 for i in range(len(Result['beam_rad'])) ]
Result2['sky_diff_rad'] = [ Result['diff_rad'][i]/3.6 for i in range(len(Result['diff_rad'])) ]
Result2['ground_diff_rad'] = [ Result['grnd_rad'][i]/3.6 for i in range(len(Result['grnd_rad'])) ]
Result2['solar_zenith'] = Result['solar_zenith']
Result2['solar_azimuth'] = Result['solar_azimuth']
new_aoi = []
new_longi = []
new_trans = []
for i in range(len(Result['aoi'])):
    if Result['aoi'][i] == None:
        new_aoi.append('None')
        new_longi.append('None')
        new_trans.append('None')
    else:
        new_aoi.append(Result['aoi'][i]*180/pi)
        new_longi.append(Result['longitudinal_angle'][i]*180/pi)
        new_trans.append(Result['transverse_angle'][i]*180/pi)
Result2['angle_of_incidence'] = new_aoi
Result2['longitudinal_angle'] = new_longi
Result2['transverse_angle'] = new_trans
Result2['system_maximum_iterations_reached'] = Result['maxIt_system']
Result2['tank_maximum_iterations_reached'] = Result['maxIt_tank']
Result2['solar_field_maximum_iterations_reached'] = Result['maxIt_solar']

Data = pd.DataFrame.from_dict(Result2)
Data.to_csv('Resultados_Detalle.csv')

os.chdir(motherFolder)
